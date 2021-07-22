import openpyxl as xl
import pandas as pd
import urllib.request
import urllib.parse
import datetime
import csv
import time
import json
import requests
from config import *

# 아파트 매매 실거래가 데이터프레임
wb1 = xl.load_workbook('아파트(매매)_실거래가_20210129105106.xlsx', data_only=True)
print(wb1.get_sheet_names())
office=wb1['아파트 매매 실거래가']

apt=[]
for row in office.rows:
    temp=[]
    for cell in row:
        temp.append(cell.value)
    apt.append(temp)
# print(apt[16:30])

df1=pd.DataFrame(apt[17:], columns=apt[16])
# print(df1)

# 연립/다세대 매매 실거래가 데이터프레임
wb2 = xl.load_workbook('연립다세대(매매)_실거래가_20210129105203.xlsx', data_only=True)
print(wb2.get_sheet_names())
office=wb2['연립다세대 매매 실거래가']

mul=[]
for row in office.rows:
    temp=[]
    for cell in row:
        temp.append(cell.value)
    mul.append(temp)
# print(mul[16:30])

df2=pd.DataFrame(mul[17:], columns=mul[16])
# print(df2)

# # 단독/다가구  매매 실거래가 데이터프레임
# wb3 = xl.load_workbook('단독다가구(매매)_실거래가_20210129105249.xlsx', data_only=True)
# print(wb3.get_sheet_names())
# detached=wb3['단독다가구 매매 실거래가']
#
# det=[]
# for row in detached.rows:
#     temp=[]
#     for cell in row:
#         temp.append(cell.value)
#     det.append(temp)
# print(det[16:30])
#
# df3=pd.DataFrame(det[17:], columns=det[16])
# print(df3)

# 오피스텔 매매 실거래가 데이터프레임
wb4 = xl.load_workbook('오피스텔(매매)_실거래가_20210128.xlsx', data_only=True)
print(wb4.get_sheet_names())
office=wb4['오피스텔 매매 실거래가']

oft=[]
for row in office.rows:
    temp=[]
    for cell in row:
        temp.append(cell.value)
    oft.append(temp)
# print(oft[16:30])

df4=pd.DataFrame(oft[17:], columns=oft[16])
# print(df4)

# url
def get_request_url(url, enc='utf-8'):
    req = urllib.request.Request(url)
    try :
        response = urllib.request.urlopen(req)
        if response.getcode() == 200:
            try:
                rcv = response.read()
                ret = rcv.decode(enc)
            except UnicodeDecodeError:
                ret = rcv.decode(enc,'replace')
            return ret
    except Exception as e:
        print(e)
        print("[%s] Error for URL: %s" % (datetime.datetime.now(), url))
        return None

# 좌표출력 오픈API
def getLocation(location):
    endPoint = 'http://api.vworld.kr/req/address?service=address&request=getcoord&version=2.0&crs=epsg:4326'
    parm = '&address='+urllib.parse.quote(location)
    parm += '&refine=false&simple=false&format=json&type=parcel'
    parm += '&key=A32C414E-511C-38A4-B53C-215AE9AC3B0C'
    url = endPoint + parm
    print(url)
    retData = get_request_url(url)
    if retData == None:
        return None
    else:
        return json.loads(retData)

# 오피스텔
lat = []
long = []
df4['소재지번'] = df4['시군구'] + df4['번지']
for item in df4['소재지번']:
    loca_json = getLocation(item)
    #     print(loca_json['response']['result']['point']['y'])
    lat.append(loca_json['response']['result']['point']['y'])
    long.append(loca_json['response']['result']['point']['x'])

df4['주택종류'] = '오피스텔'
df4['건물명'] = df4['단지명']
df4['위도'] = lat
df4['경도'] = long
df__4 = df4[['주택종류', '소재지번', '도로명', '건물명', '층', '전용면적(㎡)', '계약년월', '계약일', '거래금액(만원)', '건축년도', '위도', '경도']]

# 아파트
lat = []
long = []
df1['소재지번'] = df1['시군구'] + df1['번지']
for item in df1['소재지번']:
    loca_json = getLocation(item)
    #     print(loca_json['response']['result']['point']['y'])
    lat.append(loca_json['response']['result']['point']['y'])
    long.append(loca_json['response']['result']['point']['x'])

df1['주택종류'] = '아파트'
df1['건물명'] = df1['단지명']
df1['위도'] = lat
df1['경도'] = long
df__1 = df1[['주택종류', '소재지번', '도로명', '건물명', '층', '전용면적(㎡)', '계약년월', '계약일', '거래금액(만원)', '건축년도', '위도', '경도']]

# 연립다세대
lat=[]
long=[]
df1['소재지번']=df2['시군구']+df2['번지']
for item in df2['소재지번']:
    loca_json=getLocation(item)
#     print(loca_json['response']['result']['point']['y'])
    lat.append(loca_json['response']['result']['point']['y'])
    long.append(loca_json['response']['result']['point']['x'])

df2['주택종류']='연립/다세대'
df2['위도']=lat
df2['경도']=long
df__2=df2[['주택종류','소재지번','도로명','건물명','층','전용면적(㎡)','계약년월','계약일','거래금액(만원)','건축년도','위도','경도']]

# 데이터프레임 병합
data_all=pd.merge(df__4,df__1,df__2, how='outer')
print(data_all)

data_all.to_csv('data/서울지역 주택 매매 실거래가.csv',encoding='cp949',index=False)







